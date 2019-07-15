from openpyxl import load_workbook
import os 
import numpy as np
data_base_dir = "/home/zack/total/python_projects/ReadXlsx/Input"  # 输入文件夹的路径
outfile_dir = "/home/zack/total/python_projects/ReadXlsx/Output"  # 输出文件夹的路径
processed_number = 0  # 统计处理文件的数目
body_data = []
error_message = ''
ignore_list = ['060806', '060907']
# 检查数据是否存在错误信息
def check_data(sheet, mean, var):
    error_names = ''
    whether_wrong = True
    values = [sheet.cell(row=r, column=2).value for r in range(2, 24)]
    # 如果不在3个sigma范围内则设为False，标志为有问题
    value_check = [(abs(values[i] - mean[i]) < 3* std[i]) for i in range(len(values))]
    keys = ['Age', 'Gender', 'InterASIS', 'Bi-iliac width', 'Bi-trochanteric width', 
            'Thigh Length_L', 'Thigh Length_R', 'Calf Length_L', 'Calf Lenght_R', 'Knee Width_L', 'Knee Width_R',
            'Malleolos Width_L', 'Malleolus Width_R', 'Malleolus Height_L', 'Malleolus Height_R',
            'Foot Length_L', 'Foot Length_R', 'Foot Width_L', 'Foot Width_R', 'Height', 'Weight', 'React Time']
    # 
    for i in range(0, len(value_check)):
        if value_check[i] is False:
            error_names += keys[i] + ":{} ".format(sheet.cell(row=i+2, column=2).value)
            whether_wrong = False
    return whether_wrong, error_names

for file in sorted(os.listdir(data_base_dir)):  # 按照顺序遍历目标文件夹
    filename = os.path.join(data_base_dir, file)
    serial_number = file[0:-5]
    # print("Processing {0}".format(file))
    workbook = load_workbook(filename)
    #booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    # 获取所有的数据，组成一个列表
    if serial_number == booksheet.cell(row=1, column=2).value:
        values = [booksheet.cell(row=r, column=2).value for r in range(2, 24)]
        body_data.append(values)

# 求得均值和标准差
data_array = np.array(body_data, dtype=float)
mean = np.mean(data_array, axis=0).tolist()
std = np.std(data_array, axis=0).tolist()
print(mean)
print(std)
for file in sorted(os.listdir(data_base_dir)):  # 遍历目标文件夹
    filename = os.path.join(data_base_dir, file)
    serial_number = file[0:-5]
    # print("Processing {0}".format(file))
    workbook = load_workbook(filename)
    #booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    if serial_number == booksheet.cell(row=1, column=2).value:
        whether_wrong, error_names = check_data(booksheet, mean, std)
        if whether_wrong is False:
            error_message += "Data warning in {0} : {1}\n".format(file, error_names)
        processed_number += 1
    else:
        error_message += "File Code Error in {0}\n".format(file)

print("{0} files are processed! ".format(processed_number))
print(error_message)
