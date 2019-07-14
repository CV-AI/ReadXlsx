from openpyxl import load_workbook
import os 
import json
data_base_dir = "/home/zack/total/python_projects/ReadXlsx/Input"  # 输入文件夹的路径
outfile_dir = "/home/zack/total/python_projects/ReadXlsx/Output"  # 输出文件夹的路径
processed_number = 0  # 统计处理文件的数目
body_data = []
error_message = ''

# 检查数据是否存在错误信息
def check_data(sheet):
    error_message = ''
    checked = True
    thigh_length_checked = abs(sheet.cell(row=7, column=2).value - sheet.cell(row=8, column=2).value) <= 30
    calf_length_checked = abs(sheet.cell(row=9, column=2).value - sheet.cell(row=10, column=2).value) <= 30
    knee_width_checked = abs(sheet.cell(row=11, column=2).value - sheet.cell(row=12, column=2).value) <= 8
    malleolus_width_checked = abs(sheet.cell(row=13, column=2).value - sheet.cell(row=14, column=2).value) <=6
    malleolus_height_checked = abs(sheet.cell(row=15, column=2).value - sheet.cell(row=16, column=2).value) <=6
    foot_length_checked = abs(sheet.cell(row=17, column=2).value - sheet.cell(row=18, column=2).value) <= 6
    foot_width_checked = abs(sheet.cell(row=19, column=2).value - sheet.cell(row=20, column=2).value) <= 5
    height_checked = 1500 < sheet.cell(row=21, column=2).value <= 1850
    react_time_checked = 0.25 < sheet.cell(row=23, column=2).value <= 0.4
    check = [thigh_length_checked, calf_length_checked, knee_width_checked, malleolus_width_checked, malleolus_height_checked,
            foot_length_checked, foot_width_checked, height_checked, react_time_checked]
    error_name = ['thigh_length ', 'calf_length ', 'knee_width ', 'malleolus_width ', 
                    'malleolus_height ', 'foot_length ', 'foot_width ', 'height ', 'react_time ']
    for i in range(len(check)):
        checked = checked and check[i]
        if check[i] is False:
            error_message += error_name[i]
    return checked, error_message

for file in os.listdir(data_base_dir):  # 遍历目标文件夹
    filename = os.path.join(data_base_dir, file)
    serial_number = file[0:-5]
    print("Processing {0}".format(file))
    workbook = load_workbook(filename)
    #booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    # #迭代所有的行
    # for row in rows:
    #     line = [col.value for col in row]
    #通过坐标读取值
    # cell_11 = booksheet.cell('A1').value      
    cell_11 = booksheet.cell(row=1, column=1).value
    if serial_number == booksheet.cell(row=1, column=2).value:
        values = [booksheet.cell(row=r, column=2).value for r in range(1, 24)]
        keys = ['Serial Number', 'Age', 'Gender', 'InterASIS', 'Bi-iliac width', 'Bi-trochanteric width', 
                'Thigh Length_L', 'Thigh Length_R', 'Calf Length_L', 'Calf Lenght_R', 'Knee Width_L', 'Knee Width_R',
                'Malleolos Width_L', 'Malleolus Width_R', 'Malleolus Height_L', 'Malleolus Height_R',
                'Foot Length_L', 'Foot Length_R', 'Foot Width_L', 'Foot Width_R', 'Height', 'Weight', 'React Time']
        data_dict = dict(zip(keys, values))
        # print(data_dict)
        body_data.append(data_dict)
        processed_number += 1
        if check_data(booksheet)[0] is False:
            error_message += "Warning!!! Data Error in {0}:{1}\n".format(file, check_data(booksheet)[1])
    else:
        error_message += "File Code Error in {0}\n".format(file)

print("{0} files are processed! ".format(processed_number))
print(error_message)
json_data = json.dumps(body_data)
with open('BodyData.json', 'w') as f:  # writing JSON object
    json.dump(json_data, f)
